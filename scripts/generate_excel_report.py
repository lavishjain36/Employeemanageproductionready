#!/usr/bin/env python3
"""
Code Analysis Report Generator
Generates Excel reports from various code analysis tools
"""

import os
import sys
import json
import xml.etree.ElementTree as ET
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import glob

def parse_jacoco_coverage(xml_file):
    """Parse JaCoCo coverage XML and return coverage data"""
    if not os.path.exists(xml_file):
        return []
    
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        coverage_data = []
        for package in root.findall('.//package'):
            package_name = package.get('name', 'Unknown')
            
            for class_elem in package.findall('.//class'):
                class_name = class_elem.get('name', 'Unknown')
                line_rate = float(class_elem.get('line-rate', 0)) * 100
                branch_rate = float(class_elem.get('branch-rate', 0)) * 100
                
                coverage_data.append({
                    'Package': package_name,
                    'Class': class_name,
                    'Line Coverage (%)': round(line_rate, 2),
                    'Branch Coverage (%)': round(branch_rate, 2),
                    'Status': 'Good' if line_rate >= 80 else 'Needs Improvement' if line_rate >= 50 else 'Poor'
                })
        
        return coverage_data
    except Exception as e:
        print(f"Error parsing JaCoCo XML: {e}")
        return []

def parse_spotbugs_results(xml_file):
    """Parse SpotBugs XML results"""
    if not os.path.exists(xml_file):
        return []
    
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        bugs = []
        for bug in root.findall('.//BugInstance'):
            bug_type = bug.get('type', 'Unknown')
            priority = bug.get('priority', 'Unknown')
            category = bug.get('category', 'Unknown')
            
            source_line = bug.find('.//SourceLine')
            file_name = source_line.get('sourcefile', 'Unknown') if source_line is not None else 'Unknown'
            line_number = source_line.get('start', 'Unknown') if source_line is not None else 'Unknown'
            
            bugs.append({
                'Bug Type': bug_type,
                'Priority': priority,
                'Category': category,
                'File': file_name,
                'Line': line_number,
                'Severity': 'High' if priority == '1' else 'Medium' if priority == '2' else 'Low'
            })
        
        return bugs
    except Exception as e:
        print(f"Error parsing SpotBugs XML: {e}")
        return []

def parse_pmd_results(xml_file):
    """Parse PMD XML results"""
    if not os.path.exists(xml_file):
        return []
    
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        violations = []
        for file_elem in root.findall('.//file'):
            file_name = file_elem.get('name', 'Unknown')
            
            for violation in file_elem.findall('.//violation'):
                rule = violation.get('rule', 'Unknown')
                priority = violation.get('priority', 'Unknown')
                line = violation.get('beginline', 'Unknown')
                message = violation.text.strip() if violation.text else 'No message'
                
                violations.append({
                    'Rule': rule,
                    'Priority': priority,
                    'File': os.path.basename(file_name),
                    'Line': line,
                    'Message': message,
                    'Severity': 'High' if priority == '1' else 'Medium' if priority == '2' else 'Low'
                })
        
        return violations
    except Exception as e:
        print(f"Error parsing PMD XML: {e}")
        return []

def parse_checkstyle_results(xml_file):
    """Parse Checkstyle XML results"""
    if not os.path.exists(xml_file):
        return []
    
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        errors = []
        for file_elem in root.findall('.//file'):
            file_name = file_elem.get('name', 'Unknown')
            
            for error in file_elem.findall('.//error'):
                line = error.get('line', 'Unknown')
                column = error.get('column', 'Unknown')
                severity = error.get('severity', 'Unknown')
                message = error.get('message', 'No message')
                source = error.get('source', 'Unknown')
                
                errors.append({
                    'Rule': source,
                    'Severity': severity,
                    'File': os.path.basename(file_name),
                    'Line': line,
                    'Column': column,
                    'Message': message
                })
        
        return errors
    except Exception as e:
        print(f"Error parsing Checkstyle XML: {e}")
        return []

def create_excel_report(coverage_data, spotbugs_data, pmd_data, checkstyle_data):
    """Create comprehensive Excel report"""
    
    # Create workbook and worksheets
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Summary data
    summary_data = [
        ['Code Analysis Report', ''],
        ['Generated on', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['Coverage Analysis', ''],
        ['Total Classes Analyzed', len(coverage_data)],
        ['Average Line Coverage (%)', round(sum(d['Line Coverage (%)'] for d in coverage_data) / len(coverage_data), 2) if coverage_data else 0],
        ['Classes with >80% Coverage', len([d for d in coverage_data if d['Line Coverage (%)'] >= 80])],
        ['', ''],
        ['Bug Analysis (SpotBugs)', ''],
        ['Total Bugs Found', len(spotbugs_data)],
        ['High Priority Bugs', len([b for b in spotbugs_data if b['Severity'] == 'High'])],
        ['Medium Priority Bugs', len([b for b in spotbugs_data if b['Severity'] == 'Medium'])],
        ['Low Priority Bugs', len([b for b in spotbugs_data if b['Severity'] == 'Low'])],
        ['', ''],
        ['Code Quality (PMD)', ''],
        ['Total Violations', len(pmd_data)],
        ['High Priority Violations', len([v for v in pmd_data if v['Severity'] == 'High'])],
        ['Medium Priority Violations', len([v for v in pmd_data if v['Severity'] == 'Medium'])],
        ['Low Priority Violations', len([v for v in pmd_data if v['Severity'] == 'Low'])],
        ['', ''],
        ['Code Style (Checkstyle)', ''],
        ['Total Style Issues', len(checkstyle_data)],
        ['Error Level Issues', len([e for e in checkstyle_data if e['Severity'] == 'error'])],
        ['Warning Level Issues', len([e for e in checkstyle_data if e['Severity'] == 'warning'])],
        ['Info Level Issues', len([e for e in checkstyle_data if e['Severity'] == 'info'])]
    ]
    
    # Add summary data to worksheet
    for row in summary_data:
        summary_ws.append(row)
    
    # Style summary sheet
    for row in summary_ws.iter_rows(min_row=1, max_row=len(summary_data)):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')
    
    # Create detailed sheets
    if coverage_data:
        coverage_df = pd.DataFrame(coverage_data)
        coverage_ws = wb.create_sheet("Coverage Details")
        for r in dataframe_to_rows(coverage_df, index=False, header=True):
            coverage_ws.append(r)
        
        # Style coverage sheet
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in coverage_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
    
    if spotbugs_data:
        bugs_df = pd.DataFrame(spotbugs_data)
        bugs_ws = wb.create_sheet("SpotBugs Details")
        for r in dataframe_to_rows(bugs_df, index=False, header=True):
            bugs_ws.append(r)
        
        # Style bugs sheet
        for cell in bugs_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    
    if pmd_data:
        pmd_df = pd.DataFrame(pmd_data)
        pmd_ws = wb.create_sheet("PMD Details")
        for r in dataframe_to_rows(pmd_df, index=False, header=True):
            pmd_ws.append(r)
        
        # Style PMD sheet
        for cell in pmd_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    
    if checkstyle_data:
        checkstyle_df = pd.DataFrame(checkstyle_data)
        checkstyle_ws = wb.create_sheet("Checkstyle Details")
        for r in dataframe_to_rows(checkstyle_df, index=False, header=True):
            checkstyle_ws.append(r)
        
        # Style checkstyle sheet
        for cell in checkstyle_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def main():
    """Main function to generate the Excel report"""
    
    # Create reports directory if it doesn't exist
    os.makedirs('reports', exist_ok=True)
    
    # Parse analysis results
    print("Parsing code analysis results...")
    
    # Coverage data
    jacoco_file = 'test-results/site/jacoco/jacoco.xml'
    coverage_data = parse_jacoco_coverage(jacoco_file)
    print(f"Found {len(coverage_data)} coverage entries")
    
    # SpotBugs data
    spotbugs_file = 'test-results/spotbugsXml.xml'
    spotbugs_data = parse_spotbugs_results(spotbugs_file)
    print(f"Found {len(spotbugs_data)} SpotBugs issues")
    
    # PMD data
    pmd_file = 'test-results/pmd.xml'
    pmd_data = parse_pmd_results(pmd_file)
    print(f"Found {len(pmd_data)} PMD violations")
    
    # Checkstyle data
    checkstyle_file = 'test-results/checkstyle-result.xml'
    checkstyle_data = parse_checkstyle_results(checkstyle_file)
    print(f"Found {len(checkstyle_data)} Checkstyle issues")
    
    # Generate Excel report
    print("Generating Excel report...")
    wb = create_excel_report(coverage_data, spotbugs_data, pmd_data, checkstyle_data)
    
    # Save the report
    report_file = 'reports/code-analysis-report.xlsx'
    wb.save(report_file)
    print(f"Excel report generated: {report_file}")
    
    # Print summary
    print("\n=== REPORT SUMMARY ===")
    print(f"Coverage: {len(coverage_data)} classes analyzed")
    print(f"SpotBugs: {len(spotbugs_data)} bugs found")
    print(f"PMD: {len(pmd_data)} violations found")
    print(f"Checkstyle: {len(checkstyle_data)} style issues found")

if __name__ == "__main__":
    main()
